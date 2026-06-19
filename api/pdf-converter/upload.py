import io
import os
import re
import tempfile
import uuid

import pandas as pd
import pdfplumber
from flask import Flask, jsonify, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

app = Flask(__name__)

MAX_UPLOAD_BYTES = 32 * 1024 * 1024
YEAR_PATTERNS = ["1st Year", "2nd Year", "3rd Year", "4th Year"]

# Student ID: any non-whitespace token (any format/length) before the award number.
LINE_RE = re.compile(
    r"^(\d{5})\s+(\d+)\s+(\S+)\s+((?:TDP|TES)-[\d\-]+)\s+(.*)$",
    re.IGNORECASE,
)


def _cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return resp


def build_xlsx_download_name(pdf_filename: str) -> str:
    base = re.sub(r"\.pdf$", "", pdf_filename or "", flags=re.IGNORECASE).strip() or "converted"
    return f"{base}-SRMS.xlsx"


def extract_rows_from_pdf_bytes(pdf_bytes: bytes):
    rows = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for raw_line in text.split("\n"):
                line = raw_line.strip()

                if (
                    not line
                    or "MARINDUQUE STATE UNIVERSITY" in line
                    or "ACCOUNTING UNIT" in line
                    or "LIST OF VALIDATED GRANTEES" in line
                    or "SEQ NO" in line
                    or "FULLNAME" in line
                    or "Batch No." in line
                    or "BOAC CAMPUS" in line
                ):
                    continue

                if not re.match(r"^\d{5}\s+\d+", line):
                    continue

                seq_match = LINE_RE.match(line)
                if not seq_match:
                    continue

                seq_no = seq_match.group(1)
                index_no = seq_match.group(2)
                student_id = seq_match.group(3).strip()
                award_number = seq_match.group(4)
                remaining = seq_match.group(5).strip()

                year_level = ""
                for yp in YEAR_PATTERNS:
                    if remaining.endswith(yp):
                        year_level = yp
                        remaining = remaining[: -len(yp)].strip()
                        break

                if not year_level:
                    continue

                parts = remaining.split()
                if len(parts) < 2:
                    continue

                program = parts[-1]
                fullname = " ".join(parts[:-1])

                rows.append(
                    {
                        "SEQ NO": seq_no,
                        "INDEX": index_no,
                        "STUDENT ID": student_id,
                        "AWARD NUMBER": award_number,
                        "FULLNAME": fullname.strip(),
                        "PROGRAM": program.strip(),
                        "YEAR LEVEL": year_level.strip(),
                    }
                )
    return rows


def create_excel_bytes(rows):
    df = pd.DataFrame(rows)
    with tempfile.NamedTemporaryFile(suffix=f"_{uuid.uuid4().hex}.xlsx", delete=False) as tmp:
        temp_path = tmp.name

    try:
        df.to_excel(temp_path, index=False)

        wb = load_workbook(temp_path)
        ws = wb.active
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        widths = {"A": 12, "B": 10, "C": 18, "D": 35, "E": 40, "F": 18, "G": 15}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        wb.save(temp_path)
        with open(temp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass


@app.route("/", methods=["OPTIONS"])
@app.route("/api/pdf-converter/upload", methods=["OPTIONS"])
@app.route("/api/pdf-converter/upload/", methods=["OPTIONS"])
def upload_options():
    r = app.make_response("")
    r.status_code = 204
    return _cors(r)


@app.route("/", methods=["POST"])
@app.route("/api/pdf-converter/upload", methods=["POST"])
@app.route("/api/pdf-converter/upload/", methods=["POST"])
def upload():
    file = request.files.get("pdf")
    if not file or file.filename == "":
        return _cors(jsonify({"error": "No file selected."})), 400

    if not file.filename.lower().endswith(".pdf"):
        return _cors(jsonify({"error": "Only .pdf files are accepted."})), 400

    pdf_bytes = file.read()
    if not pdf_bytes:
        return _cors(jsonify({"error": "Empty file."})), 400
    if len(pdf_bytes) > MAX_UPLOAD_BYTES:
        return _cors(jsonify({"error": f"File too large (max {MAX_UPLOAD_BYTES // (1024 * 1024)} MB)."})), 413

    rows = extract_rows_from_pdf_bytes(pdf_bytes)
    if not rows:
        return (
            _cors(
                jsonify(
                    {
                        "error": "No rows detected from PDF.",
                        "hint": "Expected lines like: 5-digit SEQ, index, student ID, TDP/TES award number, name, program, year level.",
                    }
                )
            ),
            422,
        )

    out_bytes = create_excel_bytes(rows)
    out = io.BytesIO(out_bytes)
    out.seek(0)

    resp = send_file(
        out,
        as_attachment=True,
        download_name=build_xlsx_download_name(file.filename),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["X-Rows-Extracted"] = str(len(rows))
    return _cors(resp)
