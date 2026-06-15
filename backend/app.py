"""
SRMS PDF → Excel backend.

Run:
  python -m pip install -r backend/requirements.txt
  cd backend && python app.py

Routes:
  GET  /              — upload UI (templates/index.html)
  POST /upload        — multipart field "pdf" → styled .xlsx (grantee line parser)
  POST /convert       — multipart field "file" → .xlsx (raw pdfplumber tables)
  GET  /health        — JSON liveness
"""

from __future__ import annotations

import io
import os
import re
import uuid
from pathlib import Path

import pandas as pd
import pdfplumber
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename

app = Flask(__name__)

MAX_UPLOAD_BYTES = 32 * 1024 * 1024

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = str(BASE_DIR / "uploads")
OUTPUT_FOLDER = str(BASE_DIR / "outputs")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["OUTPUT_FOLDER"] = OUTPUT_FOLDER


# --------------------------------------------------
# Extract rows from PDF (text + regex — MSU-style lines)
# --------------------------------------------------

YEAR_PATTERNS = ["1st Year", "2nd Year", "3rd Year", "4th Year"]

LINE_RE = re.compile(
    r"^(\d{5})\s+(\d+)\s+([A-Z0-9]+)\s+((?:TDP|TES)-[\d\-]+)\s+(.*)$",
)


def extract_rows_from_pdf(pdf_path: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for raw_line in text.split("\n"):
                line = raw_line.strip()

                if (
                    not line
                    or "MARINDUQUE STATE UNIVERSITY" in line
                    or "ACCOUNTING UNIT" in line
                    or "LIST OF VALIDATED GRANTEES" in line
                    or "SEQ NO" in line
                    or "FULLNAME" in line
                    or "Batch No." in line
                    or "BOAC CAMPUS" in line
                ):
                    continue

                if not re.match(r"^\d{5}\s+\d+", line):
                    continue

                seq_match = LINE_RE.match(line)
                if not seq_match:
                    continue

                seq_no = seq_match.group(1)
                index_no = seq_match.group(2)
                student_id = seq_match.group(3)
                award_number = seq_match.group(4)
                remaining = seq_match.group(5).strip()

                year_level = ""
                for yp in YEAR_PATTERNS:
                    if remaining.endswith(yp):
                        year_level = yp
                        remaining = remaining[: -len(yp)].strip()
                        break

                if not year_level:
                    continue

                parts = remaining.split()
                if len(parts) < 2:
                    continue

                program = parts[-1]
                fullname = " ".join(parts[:-1])

                rows.append(
                    {
                        "SEQ NO": seq_no,
                        "INDEX": index_no,
                        "STUDENT ID": student_id,
                        "AWARD NUMBER": award_number,
                        "FULLNAME": fullname.strip(),
                        "PROGRAM": program.strip(),
                        "YEAR LEVEL": year_level.strip(),
                    }
                )

    return rows


# --------------------------------------------------
# Create styled Excel
# --------------------------------------------------

def create_excel(rows: list[dict[str, str]], output_path: str) -> None:
    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = {"A": 12, "B": 10, "C": 18, "D": 35, "E": 40, "F": 18, "G": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)


# --------------------------------------------------
# Table-based PDF → Excel (fallback / generic)
# --------------------------------------------------

def _dedupe_columns(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, raw in enumerate(headers):
        base = (raw or "").strip() or f"Column_{i}"
        n = seen.get(base, 0)
        seen[base] = n + 1
        out.append(base if n == 0 else f"{base}_{n + 1}")
    return out


def _normalize_table(table: list[list[str | None]]) -> pd.DataFrame | None:
    if not table:
        return None
    norm = [[("" if c is None else str(c)).strip() for c in row] for row in table]
    if not any(any(cell for cell in row) for row in norm):
        return None
    width = max(len(r) for r in norm)
    padded = [r + [""] * (width - len(r)) for r in norm]
    header, body = padded[0], padded[1:]
    if not body:
        return None
    cols = _dedupe_columns(header)
    return pd.DataFrame(body, columns=cols)


def _safe_sheet_name(name: str, used: set[str]) -> str:
    name = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "Sheet"
    name = name[:31]
    base = name
    n = 1
    while name in used:
        suffix = f"_{n}"
        name = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(name)
    return name


def _pdf_tables_to_workbook(pdf_bytes: bytes) -> tuple[dict[str, pd.DataFrame], int]:
    sheets: dict[str, pd.DataFrame] = {}
    used_names: set[str] = set()
    count = 0
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pnum, page in enumerate(pdf.pages, start=1):
            for tnum, table in enumerate(page.extract_tables() or [], start=1):
                df = _normalize_table(table)
                if df is None or df.empty:
                    continue
                count += 1
                name = _safe_sheet_name(f"p{pnum}_t{tnum}", used_names)
                sheets[name] = df
    return sheets, count


def _cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp


# --------------------------------------------------
# Routes
# --------------------------------------------------

@app.get("/health")
def health():
    return jsonify({"ok": True, "service": "pdf-to-excel"})


@app.get("/")
def home():
    return render_template("index.html")


def build_xlsx_download_name(pdf_filename: str) -> str:
    base = re.sub(r"\.pdf$", "", pdf_filename or "", flags=re.IGNORECASE).strip() or "converted"
    return f"{base}-SRMS.xlsx"


@app.route("/upload", methods=["OPTIONS"])
def upload_options():
    r = app.make_response("")
    r.status_code = 204
    r.headers["Access-Control-Allow-Origin"] = "*"
    r.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return r


@app.post("/upload")
def upload_pdf():
    file = request.files.get("pdf")
    if not file or file.filename == "":
        return _cors(jsonify({"error": "No file selected."})), 400

    if not file.filename.lower().endswith(".pdf"):
        return _cors(jsonify({"error": "Only .pdf files are accepted."})), 400

    safe_name = secure_filename(file.filename) or "upload.pdf"
    pdf_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uuid.uuid4().hex}_{safe_name}")

    file.save(pdf_path)
    try:
        if os.path.getsize(pdf_path) > MAX_UPLOAD_BYTES:
            return _cors(jsonify({"error": f"File too large (max {MAX_UPLOAD_BYTES // (1024 * 1024)} MB)."})), 413

        rows = extract_rows_from_pdf(pdf_path)
        if not rows:
            return (
                _cors(
                    jsonify(
                        {
                            "error": "No rows detected from PDF.",
                            "hint": "Expected lines like: 5-digit SEQ, index, student ID, TDP-… award, name, program, year level.",
                        }
                    )
                ),
                422,
            )

        out_name = f"converted_{uuid.uuid4().hex}.xlsx"
        output_excel = os.path.join(app.config["OUTPUT_FOLDER"], out_name)
        create_excel(rows, output_excel)

        resp = send_file(
            output_excel,
            as_attachment=True,
            download_name=build_xlsx_download_name(file.filename),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp.headers["X-Rows-Extracted"] = str(len(rows))
        return _cors(resp)
    finally:
        try:
            os.remove(pdf_path)
        except OSError:
            pass


@app.route("/convert", methods=["OPTIONS"])
def convert_options():
    r = app.make_response("")
    r.status_code = 204
    r.headers["Access-Control-Allow-Origin"] = "*"
    r.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return r


@app.post("/convert")
def convert():
    if "file" not in request.files:
        return _cors(jsonify({"error": 'Missing form field "file" (PDF upload).'})), 400
    upload = request.files["file"]
    if not upload.filename:
        return _cors(jsonify({"error": "Empty filename."})), 400
    if not upload.filename.lower().endswith(".pdf"):
        return _cors(jsonify({"error": "Only .pdf files are accepted."})), 400

    data = upload.read()
    if not data:
        return _cors(jsonify({"error": "Empty file."})), 400
    if len(data) > MAX_UPLOAD_BYTES:
        return _cors(jsonify({"error": f"File too large (max {MAX_UPLOAD_BYTES // (1024 * 1024)} MB)."})), 413

    try:
        sheets, n_tables = _pdf_tables_to_workbook(data)
    except Exception as exc:  # noqa: BLE001
        return _cors(jsonify({"error": "Could not read PDF.", "detail": str(exc)})), 400

    if not sheets:
        return (
            _cors(
                jsonify(
                    {
                        "error": "No tables found in PDF.",
                        "hint": "Try a text-based PDF; scanned images need OCR first.",
                    }
                )
            ),
            422,
        )

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    out.seek(0)

    stem = Path(upload.filename).stem or "converted"
    resp = send_file(
        out,
        as_attachment=True,
        download_name=f"{stem}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["X-Tables-Extracted"] = str(n_tables)
    return _cors(resp)


if __name__ == "__main__":
    # PDF service runs separately from the Node API (port 5000) to avoid route conflicts.
    app.run(host="127.0.0.1", port=5001, debug=True)
